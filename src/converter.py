# MIT License
#
# Copyright (c) 2025 Florian Lorenzen
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

"""
Converter Module for Friesen Enrollment Converter

This module converts Excel enrollment data to filled PDF forms.
It processes each row in the Excel file and generates a PDF page with
the enrollment data filled into a form template.
"""

import os
import sys
import tempfile
import shutil
from datetime import datetime
from typing import Dict, List, Optional, Callable, Any
from pathlib import Path

try:
    import openpyxl
    from pypdf import PdfReader, PdfWriter
    from reportlab.lib.colors import black, toColor
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
except ImportError as e:
    # Handle missing dependencies gracefully
    missing_deps = []
    if "openpyxl" in str(e):
        missing_deps.append("openpyxl")
    if "pypdf" in str(e):
        missing_deps.append("pypdf")
    if "reportlab" in str(e):
        missing_deps.append("reportlab")
    
    print(f"Missing dependencies: {', '.join(missing_deps)}")
    print("Please install with: pip install openpyxl pypdf reportlab")


class ConversionError(Exception):
    """Custom exception for conversion errors"""
    pass


class Converter:
    """
    Converts Excel enrollment data to filled PDF forms.
    
    This class reads Excel files containing enrollment data and generates
    PDF documents with one page per enrollment record.
    """

    def __init__(self, progress_callback: Optional[Callable[[str], None]] = None):
        """
        Initialize the Converter.

        Args:
            progress_callback: Optional callback function to report progress updates
        """
        self.progress_callback = progress_callback
        self.temp_dir = None

    def _report_progress(self, message: str):
        """Report progress to callback if available"""
        if self.progress_callback:
            self.progress_callback(message)
        else:
            print(message)

    def _get_form_template_path(self) -> Optional[Path]:
        """Get the path to the PDF form template"""
        # Try to find the form template in different locations
        possible_paths = [
            # When running from source
            Path(__file__).parent.parent / "resources" / "form.pdf",
            Path("resources") / "form.pdf",
            Path("form.pdf"),
        ]
        
        # When running from PyInstaller bundle
        if hasattr(sys, '_MEIPASS'):
            # PyInstaller creates a temp folder and stores path in _MEIPASS
            meipass_path = getattr(sys, '_MEIPASS', None)
            if meipass_path:
                bundle_resource_path = Path(meipass_path) / "resources" / "form.pdf"
                possible_paths.insert(0, bundle_resource_path)
        
        for path in possible_paths:
            if path.exists():
                return path
        
        # If no template found, we'll create a basic form
        return None

    def _read_excel_data(self, excel_file: str) -> List[Dict[str, Any]]:
        """
        Read the Excel file and return the data as a list of dictionaries.

        Args:
            excel_file: Path to the Excel file

        Returns:
            List of dictionaries containing the enrollment data

        Raises:
            ConversionError: If the Excel file cannot be read
        """
        try:
            # Load the workbook and get the active worksheet
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            worksheet = workbook.active
            
            if worksheet is None:
                raise ConversionError("No active worksheet found in Excel file")
            
            # Get the header row (first row)
            headers = []
            for cell in worksheet[1]:
                headers.append(cell.value if cell.value is not None else "")
            
            # Read all data rows
            data = []
            max_row = worksheet.max_row
            if max_row and max_row > 1:
                for row_num in range(2, max_row + 1):
                    row_data = {}
                    for col_num, header in enumerate(headers, 1):
                        cell_value = worksheet.cell(row=row_num, column=col_num).value
                        # Convert None to empty string, keep other values as-is
                        row_data[header] = "" if cell_value is None else str(cell_value)
                    
                    # Only add row if it has some data (not all empty)
                    if any(value.strip() for value in row_data.values() if isinstance(value, str)):
                        data.append(row_data)
            
            workbook.close()
            self._report_progress(f"Successfully read {len(data)} rows from Excel file")
            return data
            
        except Exception as e:
            raise ConversionError(f"Error reading Excel file: {e}")

    def _create_filled_pdf_page(self, row_data: Dict, page_number: int, template_path: Optional[Path] = None) -> str:
        """
        Create a filled PDF page for a single row of data.

        Args:
            row_data: Dictionary containing the data for one row
            page_number: Page number for file naming
            template_path: Path to the PDF template (optional)

        Returns:
            Path to the created PDF file
        """
        if self.temp_dir is None:
            raise ConversionError("Temporary directory not initialized")
        filename = os.path.join(self.temp_dir, f"filled_form_{page_number}.pdf")

        if template_path and template_path.exists():
            # TODO: Implement PDF form filling with existing template
            # For now, create a basic form (like the original)
            self._create_basic_form_page(row_data, filename)
        else:
            # Create basic form using reportlab
            self._create_basic_form_page(row_data, filename)

        return filename

    def _create_basic_form_page(self, row_data: Dict, filename: str):
        """
        Create a basic PDF form page using reportlab.

        Args:
            row_data: Dictionary containing the data for one row
            filename: Output filename for the PDF
        """
        # Create a canvas
        c = canvas.Canvas(filename, pagesize=A4)
        width, height = A4

        # Title
        c.setFont("Helvetica-Bold", 24)
        title = "Anmeldeformular"
        title_width = c.stringWidth(title, "Helvetica-Bold", 24)
        c.drawString((width - title_width) / 2, height - 80, title)

        # Current date
        c.setFont("Helvetica", 10)
        current_date = datetime.now().strftime("%d.%m.%Y")
        date_text = f"Datum: {current_date}"
        c.drawString(width - 150, height - 60, date_text)

        # Form fields setup
        y_start = height - 150
        field_height = 25
        field_width = 200
        label_width = 150
        spacing = 60

        # Field definitions with data mapping
        fields = [
            {
                "name": "nachname",
                "label": "Nachname:",
                "required": True,
                "value": str(row_data.get("Nachname", "")),
            },
            {
                "name": "vorname",
                "label": "Vorname:",
                "required": True,
                "value": str(row_data.get("Vorname", "")),
            },
            {
                "name": "geburtsdatum",
                "label": "Geburtsdatum:",
                "required": True,
                "value": str(row_data.get("Geburtsdatum", "")),
            },
            {
                "name": "unterricht",
                "label": "Unterricht:",
                "required": False,
                "value": str(row_data.get("Kurs", "")),
            },
        ]

        # Draw fields with data
        for i, field in enumerate(fields):
            y_pos = y_start - (i * spacing)

            # Draw label
            c.setFont("Helvetica-Bold", 12)
            label_text = field["label"]
            if field["required"]:
                label_text += " *"
            c.drawString(50, y_pos + 5, label_text)

            # Draw field border
            field_x = 50 + label_width
            c.setStrokeColor(black)
            c.setLineWidth(1)
            c.rect(field_x, y_pos, field_width, field_height)

            # Fill in the data
            if field["value"] and field["value"] != "nan":
                c.setFont("Helvetica", 11)
                c.setFillColor(black)
                # Position text inside the field box
                text_x = field_x + 5
                text_y = y_pos + 7
                c.drawString(text_x, text_y, field["value"])

            # Add placeholder text for date field if empty
            if field["name"] == "geburtsdatum" and (
                not field["value"] or field["value"] == "nan"
            ):
                c.setFont("Helvetica", 9)
                c.setFillColor(toColor("gray"))
                c.drawString(field_x + 5, y_pos - 15, "(TT.MM.JJJJ)")
                c.setFillColor(black)

        # Add instructions
        instructions_y = y_start - (len(fields) * spacing) - 40
        c.setFont("Helvetica", 10)
        c.drawString(50, instructions_y, "* Pflichtfelder")

        # Add footer
        footer_y = 80
        c.setFont("Helvetica", 8)
        c.setFillColor(toColor("gray"))
        footer_text = "Bitte füllen Sie alle Felder aus und senden Sie das Formular zurück."
        footer_width = c.stringWidth(footer_text, "Helvetica", 8)
        c.drawString((width - footer_width) / 2, footer_y, footer_text)

        # Add signature section
        signature_y = footer_y + 60
        c.setFont("Helvetica-Bold", 12)
        c.setFillColor(black)
        c.drawString(50, signature_y, "Unterschrift:")

        # Draw signature line
        sig_line_x = 150
        sig_line_width = 200
        c.setStrokeColor(black)
        c.setLineWidth(1)
        c.line(sig_line_x, signature_y - 5, sig_line_x + sig_line_width, signature_y - 5)

        # Date signature line
        date_sig_x = sig_line_x + sig_line_width + 50
        c.drawString(date_sig_x, signature_y, "Datum:")
        date_line_x = date_sig_x + 50
        date_line_width = 100
        c.line(date_line_x, signature_y - 5, date_line_x + date_line_width, signature_y - 5)

        # Save the PDF
        c.save()

    def _combine_pdfs(self, pdf_files: List[str], output_filename: str) -> str:
        """
        Combine multiple PDF files into a single PDF.

        Args:
            pdf_files: List of PDF file paths to combine
            output_filename: Output filename for the combined PDF

        Returns:
            Path to the combined PDF file

        Raises:
            ConversionError: If PDFs cannot be combined
        """
        try:
            writer = PdfWriter()

            for pdf_file in pdf_files:
                reader = PdfReader(pdf_file)
                for page in reader.pages:
                    writer.add_page(page)

            with open(output_filename, "wb") as output_file:
                writer.write(output_file)

            self._report_progress(f"Successfully combined {len(pdf_files)} pages into PDF")
            return output_filename
        except Exception as e:
            raise ConversionError(f"Error combining PDFs: {e}")

    def _cleanup_temp_files(self):
        """Clean up temporary files created during processing."""
        if self.temp_dir and os.path.exists(self.temp_dir):
            try:
                shutil.rmtree(self.temp_dir)
                self._report_progress("Temporary files cleaned up")
            except Exception as e:
                self._report_progress(f"Warning: Could not clean up temporary files: {e}")

    def convert_excel_to_pdf(self, excel_file: str, output_file: str) -> str:
        """
        Convert an Excel file to a PDF with filled forms.

        Args:
            excel_file: Path to the input Excel file
            output_file: Path to the output PDF file

        Returns:
            Path to the created PDF file

        Raises:
            ConversionError: If conversion fails
        """
        self.temp_dir = tempfile.mkdtemp()
        
        try:
            # Check dependencies
            try:
                import openpyxl
                from pypdf import PdfReader, PdfWriter
                from reportlab.pdfgen import canvas
            except ImportError as e:
                raise ConversionError(f"Missing required dependencies: {e}")

            # Read Excel data
            self._report_progress("Reading Excel file...")
            data = self._read_excel_data(excel_file)

            if not data:
                raise ConversionError("Excel file is empty")

            # Get form template
            template_path = self._get_form_template_path()
            if template_path:
                self._report_progress(f"Using form template: {template_path}")
            else:
                self._report_progress("No form template found, creating basic forms")

            # Create filled PDF pages
            pdf_files = []
            total_rows = len(data)

            self._report_progress(f"Processing {total_rows} enrollment records...")

            for counter, row_data in enumerate(data, start=1):
                # Report progress
                name = f"{row_data.get('Nachname', 'N/A')}, {row_data.get('Vorname', 'N/A')}"
                self._report_progress(f"Processing {counter}/{total_rows}: {name}")

                pdf_file = self._create_filled_pdf_page(row_data, counter, template_path)
                pdf_files.append(pdf_file)

            # Combine all PDFs
            self._report_progress("Combining PDF pages...")
            combined_pdf = self._combine_pdfs(pdf_files, output_file)

            # Clean up temporary files
            self._cleanup_temp_files()

            self._report_progress(f"Conversion completed successfully: {output_file}")
            return combined_pdf

        except Exception as e:
            self._cleanup_temp_files()
            if isinstance(e, ConversionError):
                raise
            else:
                raise ConversionError(f"Unexpected error during conversion: {e}")

    def validate_excel_file(self, excel_file: str) -> Dict[str, Any]:
        """
        Validate an Excel file and return information about it.

        Args:
            excel_file: Path to the Excel file

        Returns:
            Dictionary with validation results and file info

        Raises:
            ConversionError: If file cannot be validated
        """
        try:
            data = self._read_excel_data(excel_file)
            
            # Get column names from first row (if data exists)
            columns = list(data[0].keys()) if data else []
            
            # Check required columns
            required_columns = ["Nachname", "Vorname"]
            optional_columns = ["Geburtsdatum", "Kurs"]
            
            missing_required = [col for col in required_columns if col not in columns]
            available_optional = [col for col in optional_columns if col in columns]
            
            return {
                "valid": len(missing_required) == 0,
                "total_rows": len(data),
                "total_columns": len(columns),
                "columns": columns,
                "missing_required": missing_required,
                "available_optional": available_optional,
                "has_data": len(data) > 0
            }
            
        except Exception as e:
            raise ConversionError(f"Error validating Excel file: {e}") 
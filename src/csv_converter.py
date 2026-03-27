# MIT License
#
# Copyright (c) 2025 Florian Lorenzen
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.
"""
Enrollment file converter: CSV (semicolon-delimited) and Excel (.xlsx) to PDF.

CSV files use Windows code pages or UTF-8 (auto-detected). Excel (.xlsx) files
use sheet "Daten_zur_Verarbeitung" when the workbook has multiple sheets;
with a single sheet, that sheet is used regardless of name. Headers are in
the first row.
"""

import csv
import sys
import re
from datetime import date, datetime
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

from openpyxl import load_workbook
from charset_normalizer import from_path
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm, mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.colors import black, grey, Color
from reportlab.pdfgen import canvas

key_mapping = {
    'id': None,
    'sorting': None,
    'form': None,
    'date': None,
    'ip': None,
    'fd_member': None,
    'fd_user': None,
    'fd_member_group': None,
    'fd_user_group': None,
    'published': None,
    'alias': None,
    'confirmationSent': None,
    'confirmationDate': None,
    'be_notes': None,
    'Angebot': 'goal',
    'von': 'from',
    'bis': 'to',
    'Beitrag': None,
    'Unterricht': None,
    'Vorkentnisse': 'previous_knowledge',
    'Geschlecht': 'sex',
    'Teilnehmer-Name': 'last_name',
    'Teilnehmer-Vorname': 'first_name',
    'Geburtsdatum': 'birth_date',
    'Strasse_Teilnehmer': 'street',
    'PLZ_Teilnehmer': 'zip',
    'Ort_Teilnehmer': 'city',
    'Telefon': 'phone',
    'E-Mail': 'email',
    'Beruf': 'profession',
    'Vereinsmitglied': 'is_member',
    'Vereinsmitglied_Mitgliedsnummer': 'member_id',
    'Ermaessigung': 'discount',
    'Familienmitglied': 'family_member',
    'Mitglieder_Haushalt': 'household_members',
    'Geschlecht_Antragsteller': 'applicant_sex',
    'Name-Antragsteller': 'applicant_last_name',
    'Vorname-Antragsteller': 'applicant_first_name',
    'Strasse_Antragsteller': 'applicant_street',
    'PLZ_Antragsteller': 'applicant_zip',
    'Ort_Antragsteller': 'applicant_city',
    'check-Sportgesundheit': None,
    'check-Datenschutz': None,
    'check-Coronaschutz': None,
    'Kontoinhaber': 'account_holder',
    'Kreditinstitut': 'bank_name',
    'IBAN': 'iban',
    'BIC': 'bic'
}

# Create reverse mapping for debug labels (mapped_key -> original_key)
reverse_key_mapping = {}
for original_key, mapped_key in key_mapping.items():
    if mapped_key is not None:
        reverse_key_mapping[mapped_key] = original_key

# Excel export: sheet "Daten_zur_Verarbeitung" (see mapping.txt)
xlsx_key_mapping = {
    'Anrede': 'sex',
    'Vorname': 'first_name',
    'Nachname': 'last_name',
    'Titel': None,
    'Firmenname': None,
    'Mobil': None,
    'Telefon': None,
    'E-Mail': None,
    'Strasse und Nr.': None,
    'Postleitzahl': None,
    'Stadt': None,
    'Land': None,
    'Geburtstag': 'birth_date',
    'Geburtsort': None,
    'Tags': None,
    'Preis': None,
    'Notizen (nur für Mitarbeiter sichtbar)': None,
    'Buchungscode': None,
    'Eingecheckt am': None,
    'Buchungen': 'goal',
    'Veranstaltung von': 'from',
    'Veranstaltung bis': 'to',
    'Kunde Anrede': 'applicant_sex',
    'Kunde Vorname': 'applicant_first_name',
    'Kunde Nachname': 'applicant_last_name',
    'Kunde Titel': None,
    'Kunde Firma': None,
    'Kunde Mobil': 'phone',
    'Kunde Telefon': None,
    'Kunde E-Mail': 'email',
    'Kunde Notiz': None,
    'Kunde Tags': None,
    'Kunde Strasse': 'street',
    'Kunde Postleitzahl': 'zip',
    'Kunde Stadt': 'city',
    'Kunde Land': None,
    'Kunde Ausstehend': None,
    'Kunde Fällig': None,
    'Kunde Gesamtbetrag': None,
    'Name d. Kontoinhabers': 'account_holder',
    # Some exports misspell "Beitragskontos" as "Beitagskontos" in the column title
    'IBAN des Beitagskontos': 'iban',
    'IBAN des Beitragskontos': 'iban',
    'BIC des Beitragskontos': 'bic',
    'Bank des Beitragskontos': 'bank_name',
}

xlsx_reverse_key_mapping: Dict[str, str] = {}
for original_key, mapped_key in xlsx_key_mapping.items():
    if mapped_key is not None:
        xlsx_reverse_key_mapping[mapped_key] = original_key

XLSX_SHEET_NAME = 'Daten_zur_Verarbeitung'

# Excel imports: fixed "von"/"bis" dates for the PDF (ignore sheet values).
XLSX_FIXED_FROM_DATE = '09.02.2026'
XLSX_FIXED_TO_DATE = '05.07.2026'

_DE_IBAN_LEN = 22


def normalize_iban(value: Any) -> str:
    """Strip whitespace and uppercase for IBAN comparison and formatting."""
    if value is None:
        return ''
    return re.sub(r'\s+', '', str(value).strip().upper())


def validate_iban(value: Any) -> Optional[str]:
    """
    Return None if empty or valid; otherwise a short English error for the GUI.
    Order: normalize, empty OK, charset, DE length 22, non-DE length 15–34, mod-97.
    """
    iban = normalize_iban(value)
    if not iban:
        return None
    if not re.fullmatch(r'[A-Z0-9]+', iban):
        return 'Invalid format (letters and digits only).'
    if len(iban) < 2:
        return 'Invalid format (country code missing).'
    country = iban[:2]
    if country == 'DE':
        if len(iban) != _DE_IBAN_LEN:
            return f'German (DE) IBAN must be 22 characters (got {len(iban)}).'
    else:
        if not (15 <= len(iban) <= 34):
            return f'Invalid IBAN length ({len(iban)} characters; allowed 15–34).'
    rearranged = iban[4:] + iban[:4]
    parts: List[str] = []
    for ch in rearranged:
        if ch.isdigit():
            parts.append(ch)
        elif 'A' <= ch <= 'Z':
            parts.append(str(ord(ch) - ord('A') + 10))
        else:
            return 'Invalid format.'
    num_str = ''.join(parts)
    try:
        if int(num_str) % 97 != 1:
            return 'Invalid check digits (mod-97 verification failed).'
    except ValueError:
        return 'Invalid format.'
    return None


# ISO 9362 BIC: 4 letters (bank) + 2 letters (country) + 2 alphanumeric (location) + optional 3 (branch)
_BIC_RE = re.compile(r'^[A-Z]{4}[A-Z]{2}[A-Z0-9]{2}([A-Z0-9]{3})?$')


def normalize_bic(value: Any) -> str:
    """Strip whitespace and uppercase for BIC comparison and display."""
    if value is None:
        return ''
    return re.sub(r'\s+', '', str(value).strip().upper())


def validate_bic(value: Any) -> Optional[str]:
    """Return None if empty or valid BIC; otherwise a short English error for the GUI."""
    bic = normalize_bic(value)
    if not bic:
        return None
    if len(bic) not in (8, 11):
        return f'Invalid BIC length ({len(bic)} characters; must be 8 or 11).'
    if not _BIC_RE.fullmatch(bic):
        return 'Invalid BIC format (4 bank letters, 2 country letters, 2 location, optional 3 branch).'
    return None


def _filter_nonempty_rows(data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Drop rows where every value is empty or whitespace-only."""
    filtered_data: List[Dict[str, Any]] = []
    for row in data:
        has_content = False
        for value in row.values():
            if value and str(value).strip():
                has_content = True
                break
        if has_content:
            filtered_data.append(row)
    print(f"📊 Filtered {len(data) - len(filtered_data)} empty/whitespace rows from {len(data)} total rows")
    return filtered_data


def _normalize_xlsx_cell(value: Any) -> str:
    """Turn openpyxl cell values into display strings for the PDF pipeline."""
    if value is None:
        return ''
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0 and value.microsecond == 0:
            return value.strftime('%d.%m.%Y')
        return value.strftime('%d.%m.%Y %H:%M')
    if isinstance(value, date):
        return value.strftime('%d.%m.%Y')
    if isinstance(value, bool):
        return 'Ja' if value else 'Nein'
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def map_anrede_to_sex(val: Any) -> str:
    """Map Excel salutation (Herr/Frau/Divers) to PDF gender labels."""
    if val is None or (isinstance(val, str) and not val.strip()):
        return ''
    s = str(val).strip()
    low = s.lower()
    mapped = {'herr': 'männlich', 'frau': 'weiblich', 'divers': 'divers'}
    return mapped.get(low, s)


def read_xlsx_to_mapped_dicts(filename: str) -> List[Dict[str, Any]]:
    """
    Read an .xlsx file and map columns to PDF field keys.

    If the workbook has a single sheet, that sheet is used regardless of its name.
    If it has multiple sheets, the sheet named XLSX_SHEET_NAME is used.
    """
    try:
        wb = load_workbook(filename, read_only=True, data_only=True)
    except FileNotFoundError:
        raise FileNotFoundError(f"File '{filename}' not found")
    except Exception as e:
        raise Exception(f"Error opening Excel file '{filename}': {e}")

    try:
        names = wb.sheetnames
        if len(names) == 1:
            sheet_name = names[0]
        elif XLSX_SHEET_NAME in names:
            sheet_name = XLSX_SHEET_NAME
        else:
            raise ValueError(
                f"Sheet '{XLSX_SHEET_NAME}' not found in '{filename}'. "
                f"Available sheets: {', '.join(names)}"
            )
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ValueError(f"Sheet '{sheet_name}' in '{filename}' is empty")

        headers: List[str] = []
        for h in header_row:
            if h is None:
                headers.append('')
            else:
                headers.append(str(h).strip())

        if not any(headers):
            raise ValueError(f"File '{filename}' has no headers in sheet '{sheet_name}'")

        data: List[Dict[str, Any]] = []
        for row in rows_iter:
            cells = list(row)
            row_dict: Dict[str, str] = {}
            for i, header in enumerate(headers):
                if not header:
                    continue
                val = cells[i] if i < len(cells) else None
                row_dict[header] = _normalize_xlsx_cell(val)

            converted_row: Dict[str, Any] = {}
            for original_key, mapped_key in xlsx_key_mapping.items():
                if mapped_key is None:
                    continue
                val = row_dict.get(original_key, '')
                if mapped_key not in converted_row:
                    converted_row[mapped_key] = val
                else:
                    existing = converted_row[mapped_key]
                    if (not existing or not str(existing).strip()) and val and str(val).strip():
                        converted_row[mapped_key] = val

            if 'sex' in converted_row:
                converted_row['sex'] = map_anrede_to_sex(converted_row.get('sex', ''))
            if 'applicant_sex' in converted_row:
                converted_row['applicant_sex'] = map_anrede_to_sex(converted_row.get('applicant_sex', ''))

            converted_row['from'] = XLSX_FIXED_FROM_DATE
            converted_row['to'] = XLSX_FIXED_TO_DATE

            data.append(converted_row)

        if not data:
            print(f"Warning: File '{filename}' contains only headers, no data rows")

        return _filter_nonempty_rows(data)
    finally:
        wb.close()


def transform_phone_number(phone: str) -> str:
    """
    Transform phone number according to specific rules.
    
    Args:
        phone (str): Original phone number
        
    Returns:
        str: Transformed phone number
    """
    if not phone or not str(phone).strip():
        return phone
    
    phone_str = str(phone).strip()
    original_phone = phone_str
    
    # Step 1: If the phone content matches a floating point number convert this to an integer
    # Regex to match full exponential float syntax with comma decimal separator: [+-]?(\d+,?\d*|,\d+)([eE][+-]?\d+)?
    float_pattern = r'^[+-]?(\d+,?\d*|,\d+)([eE][+-]?\d+)?$'
    if re.match(float_pattern, phone_str):
        try:
            # Convert comma decimal separator to period for float parsing
            phone_str_for_float = phone_str.replace(',', '.')
            float_val = float(phone_str_for_float)
            phone_str = str(int(float_val))
            print(f"Phone transformation step 1 (float to int): '{original_phone}' -> '{phone_str}'")
        except (ValueError, TypeError, OverflowError):
            # Conversion failed, keep original string
            pass
    
    # Step 2: If the phone number starts with "49" add a "+" at the front
    if phone_str.startswith("49"):
        phone_str = "+" + phone_str
        print(f"Phone transformation step 2 (add + for 49): '{original_phone}' -> '{phone_str}'")
    
    # Step 3: If the phone number starts with a digit 1..9, add a zero at the front
    if phone_str and phone_str[0] in "123456789":
        phone_str = "0" + phone_str
        print(f"Phone transformation step 3 (add 0 for 1-9): '{original_phone}' -> '{phone_str}'")
    
    if phone_str != original_phone:
        print(f"Final phone transformation: '{original_phone}' -> '{phone_str}'")
    
    return phone_str

def read_csv_to_dicts(filename: str) -> List[Dict[str, Any]]:
    """
    Read a CSV file with Windows 1252 encoding and semicolon separators,
    returning a list of dictionaries where the first row contains the keys.
    
    Args:
        filename (str): Path to the CSV file to read
        
    Returns:
        List[Dict[str, Any]]: List of dictionaries, each representing a row
                              with column headers as keys
                              
    Raises:
        FileNotFoundError: If the specified file doesn't exist
        UnicodeDecodeError: If the file cannot be decoded with Windows 1252 encoding
        Exception: For other CSV parsing errors
    """
    try:
        with open(filename, 'r', encoding='windows-1252', newline='') as csvfile:
            reader = csv.DictReader(csvfile, delimiter=';')
            return list(reader)
    except FileNotFoundError:
        raise FileNotFoundError(f"File '{filename}' not found")
    except UnicodeDecodeError as e:
        raise UnicodeDecodeError(f"Failed to decode file '{filename}' with Windows 1252 encoding: {e}")
    except Exception as e:
        raise Exception(f"Error reading CSV file '{filename}': {e}")


def generate_pdf_from_dict(
    data_dict: Dict[str, Any],
    c: canvas.Canvas,
    debug: bool = False,
    debug_reverse_mapping: Optional[Dict[str, str]] = None,
    mark_iban_invalid: bool = False,
    mark_bic_invalid: bool = False,
) -> None:
    """
    Generate a one-page PDF from a dictionary row with specific formatting.
    
    Args:
        data_dict (Dict[str, Any]): Dictionary containing the row data
        c (canvas.Canvas): Canvas object to draw on
        debug (bool): If True, show source keys as labels in top-right corner of fields
        debug_reverse_mapping: mapped_key -> source column name for debug labels (defaults to CSV mapping)
        mark_iban_invalid: If True, prefix IBAN text with !!Ungültig!! (non-empty IBAN only)
        mark_bic_invalid: If True, prefix BIC text with !!Ungültig!! (non-empty BIC only)
        
    Raises:
        Exception: If there's an error creating the PDF
    """
    try:
        rev_map = debug_reverse_mapping if debug_reverse_mapping is not None else reverse_key_mapping

        # Define a very light grey color (95% white, 5% black)
        light_grey = Color(0.95, 0.95, 0.95)
        
        def draw_debug_label(original_key: str, x: float, y: float, box_width: float, box_height: float) -> None:
            """Draw debug label in top-right corner of a field"""
            if not debug or original_key in ['sex', 'applicant_sex']:  # Skip sex fields
                return
            
            # Set red color and small font
            c.setFillColor(Color(1, 0, 0))  # Red
            c.setFont("Helvetica", 6)
            
            # Position in top-right corner with small margin
            label_x = x + box_width - 2
            label_y = y + 2  # Small margin from top
            
            # Draw the label
            c.drawRightString(label_x, label_y, original_key)
            
            # Reset to black color
            c.setFillColor(black)
        
        def draw_text_in_box(text, x, y, box_width, font_name="Helvetica", base_font_size=12, padding=5, field_key=None):
            """
            Draw text in a box with automatic font size reduction if needed.
            
            Args:
                text: Text to draw
                x, y: Position coordinates
                box_width: Width of the box
                font_name: Font family name
                base_font_size: Starting font size
                padding: Padding from box edges
                field_key: Key for debug label (optional)
            """
            # Calculate available width for text
            available_width = box_width - (2 * padding)
            
            # Start with base font size
            font_size = base_font_size
            
            # Reduce font size until text fits
            while font_size > 6:  # Minimum font size of 6pt
                c.setFont(font_name, font_size)
                text_width = c.stringWidth(text, font_name, font_size)
                if text_width <= available_width:
                    break
                font_size -= 1
            
            # Draw the text if it exists
            if text:
                c.drawString(x + padding, y, text)
            
            # Draw debug label if field_key is provided (always show, even if no data)
            if field_key and field_key in rev_map:
                original_key = rev_map[field_key]
                # Estimate box height based on font size
                box_height = font_size + 8  # Approximate height
                draw_debug_label(original_key, x, y, box_width, box_height)
        
        # Set margins
        margin_top = 1.5 * cm
        margin_bottom = 1.5 * cm
        margin_left = 2 * cm
        margin_right = 2 * cm
        
        # Calculate usable area
        usable_width = A4[0] - margin_left - margin_right
        usable_height = A4[1] - margin_top - margin_bottom
        
        # Start position (top-left of usable area)
        x = margin_left
        y = A4[1] - margin_top
        
        # Main title - centered, bold, 20pt
        c.setFont("Helvetica-Bold", 20)
        c.setFillColor(black)
        title_text = "Berliner Schwimmverein „Friesen 1895“ e. V."
        title_width = c.stringWidth(title_text, "Helvetica-Bold", 20)
        c.drawString((A4[0] - title_width) / 2, y, title_text)
        
        # Move down for subtitle
        y -= 40
        
        # Subtitle - centered, bold, 14pt
        c.setFont("Helvetica-Bold", 14)
        subtitle_text = "Aufnahmeantrag (konvertiert)"
        subtitle_width = c.stringWidth(subtitle_text, "Helvetica-Bold", 14)
        c.drawString((A4[0] - subtitle_width) / 2, y, subtitle_text)
        
        # Add three boxes at the top
        # Calculate box dimensions
        top_box_width = 4 * cm
        top_box_height = 1.2 * cm  # Same height as other grey boxes
        top_box_gap = 1 * mm  # 1mm gap between boxes
        
        # Calculate positions aligned with left margin
        # First box: "Unterrichtsziel" - aligned with left margin
        unterrichtszie_x = x  # Align with left margin
        unterrichtszie_y = y + 5  # Slightly above subtitle baseline
        
        # Draw first box with light grey fill
        c.setFillColor(light_grey)
        c.rect(unterrichtszie_x, unterrichtszie_y - top_box_height, top_box_width, top_box_height, fill=1)
        c.setFillColor(black)  # Reset to black for text
        
        # Add label for first box
        c.setFont("Helvetica", 8)
        c.drawString(unterrichtszie_x + 2, unterrichtszie_y - top_box_height + 2, "Unterrichtsziel")
        
        # Second box: "von" - below first box
        von_x = unterrichtszie_x
        von_y = unterrichtszie_y - top_box_height - top_box_gap
        
        # Draw second box with light grey fill
        c.setFillColor(light_grey)
        c.rect(von_x, von_y - top_box_height, top_box_width, top_box_height, fill=1)
        c.setFillColor(black)  # Reset to black for text
        
        # Add label for second box
        c.setFont("Helvetica", 8)
        c.drawString(von_x + 2, von_y - top_box_height + 2, "von")
        
        # Third box: "bis" - right of second box
        bis_x = von_x + top_box_width + top_box_gap
        bis_y = von_y  # Same y position as second box
        
        # Draw third box with light grey fill
        c.setFillColor(light_grey)
        c.rect(bis_x, bis_y - top_box_height, top_box_width, top_box_height, fill=1)
        c.setFillColor(black)  # Reset to black for text
        
        # Add label for third box
        c.setFont("Helvetica", 8)
        c.drawString(bis_x + 2, bis_y - top_box_height + 2, "bis")
        
        # Fourth box: "Mitgliedsnummer" - at right margin, same height as "Unterrichtsziel"
        mitgliedsnummer_x = x + usable_width - top_box_width  # Right margin minus box width
        mitgliedsnummer_y = unterrichtszie_y  # Same height as "Unterrichtsziel"
        
        # Draw fourth box with light grey fill
        c.setFillColor(light_grey)
        c.rect(mitgliedsnummer_x, mitgliedsnummer_y - top_box_height, top_box_width, top_box_height, fill=1)
        c.setFillColor(black)  # Reset to black for text
        
        # Add label for fourth box
        c.setFont("Helvetica", 8)
        c.drawString(mitgliedsnummer_x + 2, mitgliedsnummer_y - top_box_height + 2, "Mitgliedsnummer")
        
        # Fifth box: "Zahlernummer" - below fourth box
        zahlernummer_x = mitgliedsnummer_x  # Same x position as fourth box
        zahlernummer_y = mitgliedsnummer_y - top_box_height - top_box_gap  # 1mm gap below
        
        # Draw fifth box with light grey fill
        c.setFillColor(light_grey)
        c.rect(zahlernummer_x, zahlernummer_y - top_box_height, top_box_width, top_box_height, fill=1)
        c.setFillColor(black)  # Reset to black for text
        
        # Add label for fifth box
        c.setFont("Helvetica", 8)
        c.drawString(zahlernummer_x + 2, zahlernummer_y - top_box_height + 2, "Zahlernummer")
        
        # Add content to the top boxes
        c.setFont("Helvetica", 12)
        
        # First box: "goal" field into "Unterrichtsziel"
        goal = data_dict.get('goal', '')
        draw_text_in_box(goal, unterrichtszie_x, unterrichtszie_y - (top_box_height / 2) - 4, top_box_width, field_key='goal')
        
        # Second box: "from" field into "von"
        from_date = data_dict.get('from', '')
        draw_text_in_box(from_date, von_x, von_y - (top_box_height / 2) - 4, top_box_width, field_key='from')
        
        # Third box: "to" field into "bis"
        to_date = data_dict.get('to', '')
        draw_text_in_box(to_date, bis_x, bis_y - (top_box_height / 2) - 4, top_box_width, field_key='to')
        
        # Fourth box: "member_id" field into "Mitgliedsnummer"
        member_id = data_dict.get('member_id', '')
        draw_text_in_box(member_id, mitgliedsnummer_x, mitgliedsnummer_y - (top_box_height / 2) - 4, top_box_width, field_key='member_id')
        
        # Shift everything down by 2cm after subtitle
        y -= 2 * cm
        
        # Move down for section heading (increased space before)
        y -= 50
        
        # Section heading - bold, 12pt, underlined
        c.setFont("Helvetica-Bold", 12)
        section_text = "I. Mitgliedschaftsdaten"
        c.drawString(x, y, section_text)
        
        # Draw underline for section heading
        text_width = c.stringWidth(section_text, "Helvetica-Bold", 12)
        c.line(x, y - 2, x + text_width, y - 2)
        
        # Move down for the box (reduced space after)
        y -= 10
        
        # Draw the enlarged main box (now with bottom line to close it)
        c.setStrokeColor(black)
        c.setLineWidth(1)
        main_box_height = 235  # Enlarged by ~4mm (11 points)
        # Draw all four sides to close the box
        c.line(x, y, x + usable_width, y)  # Top line
        c.line(x, y, x, y - main_box_height)  # Left line
        c.line(x + usable_width, y, x + usable_width, y - main_box_height)  # Right line
        c.line(x, y - main_box_height, x + usable_width, y - main_box_height)  # Bottom line
        
        # Add gender selection text inside the box (top line)
        c.setFont("Helvetica", 10)
        c.setFillColor(black)
        
        # Get gender value from data
        sex = data_dict.get('sex', '')
        if sex is not None:
            sex = sex.lower()
        else:
            sex = ''
        
        # Create gender selection text with checkboxes
        gender_text = "Geschlecht: "
        
        # Position the text (right-aligned within the box)
        base_text = "Geschlecht: O weiblich / O männlich / O divers"
        text_width = c.stringWidth(base_text, "Helvetica", 10)
        c.drawRightString(x + usable_width - 10, y - 20, base_text)
        
        # Add checkmarks by overprinting "x" over "o" based on sex value
        if sex in ['weiblich', 'männlich', 'divers']:
            # Calculate positions for the "o" characters
            base_x = x + usable_width - 10 - text_width  # Starting position of the text
            
            # Position of each "o" in the text
            o_positions = []
            current_x = base_x
            for i, char in enumerate(base_text):
                if char == 'O':
                    o_positions.append(current_x)
                current_x += c.stringWidth(char, "Helvetica", 10)
            
            # Overprint "x" on the appropriate "o"
            if sex == 'weiblich' and len(o_positions) >= 1:
                c.drawString(o_positions[0], y - 20, "X")
            elif sex == 'männlich' and len(o_positions) >= 2:
                c.drawString(o_positions[1], y - 20, "X")
            elif sex == 'divers' and len(o_positions) >= 2:
                c.drawString(o_positions[2], y - 20, "X")
        
        # Position for the name and birth date boxes inside the main box (moved up more)
        inner_y = y - 40  # Move down inside the main box (reduced from 50 to 40)
        
        # Calculate inner box dimensions (slightly smaller to fit inside)
        inner_margin = 10  # Margin from main box edges
        inner_width = usable_width - (2 * inner_margin)
        inner_box_height = 1.2 * cm  # Slightly smaller height
        box_gap = 5  # Gap between boxes
        left_box_width = (inner_width * 2) / 3 - box_gap / 2
        right_box_width = inner_width / 3 - box_gap / 2
        
        # Draw the two adjacent boxes inside the main box with light grey fill
        c.setFillColor(light_grey)
        c.rect(x + inner_margin, inner_y - inner_box_height, left_box_width, inner_box_height, fill=1)  # Left box
        c.rect(x + inner_margin + left_box_width + box_gap, inner_y - inner_box_height, right_box_width, inner_box_height, fill=1)  # Right box
        c.setFillColor(black)  # Reset to black for text
        
        # Add labels in 8pt font in lower left corner of each inner box
        c.setFont("Helvetica", 8)
        c.drawString(x + inner_margin + 2, inner_y - inner_box_height + 2, "Name, Vorname")
        c.drawString(x + inner_margin + left_box_width + box_gap + 2, inner_y - inner_box_height + 2, "geboren am")
        
        # Add content in 12pt font, left-aligned, baseline in middle of inner boxes
        c.setFont("Helvetica", 12)
        
        # Left box: last_name, first_name
        last_name = data_dict.get('last_name', '')
        first_name = data_dict.get('first_name', '')
        name_text = f"{last_name}, {first_name}" if last_name and first_name else f"{last_name}{first_name}"
        draw_text_in_box(name_text, x + inner_margin, inner_y - (inner_box_height / 2) - 4, left_box_width, field_key='last_name')
        
        # Right box: birth_date
        birth_date = data_dict.get('birth_date', '')
        draw_text_in_box(birth_date, x + inner_margin + left_box_width + box_gap, inner_y - (inner_box_height / 2) - 4, right_box_width, field_key='birth_date')
        
        # Add another box below the name/birth date boxes
        discount_y = inner_y - inner_box_height - 10  # Position below the previous boxes with 10pt gap
        
        # Calculate dimensions for the discount box
        discount_box_height = 1.8 * cm  # Increased height from 1.2cm to 1.8cm
        discount_left_width = inner_width / 8  # Reduced from 1/6 to 1/8 of the width
        discount_right_width = inner_width - discount_left_width  # Rest of the width
        
        # Draw the discount box (no fill)
        c.rect(x + inner_margin, discount_y - discount_box_height, inner_width, discount_box_height)
        
        # Add content to the discount box
        c.setFont("Helvetica", 10)
        
        # Right column: Multi-line text
        right_text_line1 = "O ermäßigtes Mitglied, da bereits mind. 2 Mitglieder meines Haushaltes Mitglieder"
        right_text_line2 = "ohne Begrenzung der Mitgliedsdauer sind."
        right_text_line3 = "Diese sind"
        
        # Calculate text positioning for right column
        right_x = x + inner_margin + discount_left_width + 5
        line1_y = discount_y - (discount_box_height / 2) + 10  # Moved up by ~15pt
        line2_y = line1_y - 15  # 15pt gap between lines
        line3_y = line2_y - 15  # 15pt gap between lines
        
        # Left column: "als" - aligned with first line
        c.drawString(x + inner_margin + 5, line1_y, "als")
        
        c.drawString(right_x, line1_y, right_text_line1)
        c.drawString(right_x, line2_y, right_text_line2)
        c.drawString(right_x, line3_y, right_text_line3)
        
        # Add family member content in a box if defined
        family_member = data_dict.get('family_member', '')
        if family_member:
            # Calculate position of "O" in the first line (more precise)
            o_position = right_x + 1
            c.drawString(o_position, line1_y, "X")
            # Calculate position for family member box
            family_box_x = right_x + c.stringWidth(right_text_line3, "Helvetica", 10) + 5
            family_box_width = (x + inner_margin + inner_width) - family_box_x - 5  # Fixed width extending to right edge minus 5
            family_box_height = 14  # Reduced height by 6pt (from 18 to 12)
            
            # Draw the box around family member text (shifted up 9pt total, extends above and below baseline)
            c.setFillColor(light_grey)
            c.rect(family_box_x, line3_y - 3, family_box_width, family_box_height, fill=1)
            c.setFillColor(black)  # Reset to black for text
            
            # Add family member text inside the box (same baseline as "Diese sind")
            draw_text_in_box(family_member, family_box_x, line3_y+1, family_box_width, padding=2, field_key='family_member')
        
        
        # Add address row below the discount box
        address_y = discount_y - discount_box_height - 10  # Position below discount box with 10pt gap
        
        # Calculate dimensions for address boxes
        address_box_height = 1.2 * cm
        address_box_gap = 5  # Gap between address boxes
        address_box1_width = (inner_width * 0.6) - address_box_gap  # 60% of width minus gap
        address_box2_width = (inner_width * 0.15) - address_box_gap  # 15% of width minus gap
        address_box3_width = (inner_width * 0.25)  # 25% of width minus gap plus fixed amount
        
        # Draw the three address boxes with light grey fill
        c.setFillColor(light_grey)
        c.rect(x + inner_margin, address_y - address_box_height, address_box1_width, address_box_height, fill=1)  # Street
        c.rect(x + inner_margin + address_box1_width + address_box_gap, address_y - address_box_height, address_box2_width, address_box_height, fill=1)  # ZIP
        c.rect(x + inner_margin + address_box1_width + address_box_gap + address_box2_width + address_box_gap, address_y - address_box_height, address_box3_width, address_box_height, fill=1)  # City
        c.setFillColor(black)  # Reset to black for text
        
        # Add labels in 8pt font in lower left corner of each box
        c.setFont("Helvetica", 8)
        c.drawString(x + inner_margin + 2, address_y - address_box_height + 2, "Straße, Hausnummer")
        c.drawString(x + inner_margin + address_box1_width + address_box_gap + 2, address_y - address_box_height + 2, "Postleitzahl")
        c.drawString(x + inner_margin + address_box1_width + address_box_gap + address_box2_width + address_box_gap + 2, address_y - address_box_height + 2, "Ort")
        
        # Add content in 12pt font, left-aligned, baseline in middle of boxes
        c.setFont("Helvetica", 12)
        
        # First box: street
        street = data_dict.get('street', '')
        draw_text_in_box(street, x + inner_margin, address_y - (address_box_height / 2) - 4, address_box1_width, field_key='street')
        
        # Second box: zip
        zip_code = data_dict.get('zip', '')
        draw_text_in_box(zip_code, x + inner_margin + address_box1_width + address_box_gap, address_y - (address_box_height / 2) - 4, address_box2_width, field_key='zip')
        
        # Third box: city
        city = data_dict.get('city', '')
        draw_text_in_box(city, x + inner_margin + address_box1_width + address_box_gap + address_box2_width + address_box_gap, address_y - (address_box_height / 2) - 4, address_box3_width, field_key='city')
        
        # Add contact information row below the address boxes
        contact_y = address_y - address_box_height - 10  # Position below address boxes with 10pt gap
        
        # Calculate dimensions for contact boxes (email wider, phone smaller)
        contact_box_height = 1.2 * cm
        contact_box_gap = 5  # Gap between contact boxes
        base_contact_box_width = (inner_width / 3) - (2 * contact_box_gap / 3)  # Base 1/3 width minus gap
        
        # Adjust widths: email +1cm, phone -1cm from left
        email_box_width = base_contact_box_width + 1 * cm
        phone_box_width = base_contact_box_width - 1 * cm
        profession_box_width = base_contact_box_width  # Keep profession box unchanged
        
        # Draw the three contact boxes with light grey fill
        c.setFillColor(light_grey)
        c.rect(x + inner_margin, contact_y - contact_box_height, email_box_width, contact_box_height, fill=1)  # Email
        c.rect(x + inner_margin + email_box_width + contact_box_gap, contact_y - contact_box_height, phone_box_width, contact_box_height, fill=1)  # Phone
        c.rect(x + inner_margin + email_box_width + contact_box_gap + phone_box_width + contact_box_gap, contact_y - contact_box_height, profession_box_width, contact_box_height, fill=1)  # Profession
        c.setFillColor(black)  # Reset to black for text
        
        # Add labels in 8pt font in lower left corner of each box
        c.setFont("Helvetica", 8)
        c.drawString(x + inner_margin + 2, contact_y - contact_box_height + 2, "E-Mail")
        c.drawString(x + inner_margin + email_box_width + contact_box_gap + 2, contact_y - contact_box_height + 2, "Telefon / Mobil")
        c.drawString(x + inner_margin + email_box_width + contact_box_gap + phone_box_width + contact_box_gap + 2, contact_y - contact_box_height + 2, "Beruf")
        
        # Add content in 12pt font, left-aligned, baseline in middle of boxes
        c.setFont("Helvetica", 12)
        
        # First box: email
        email = data_dict.get('email', '')
        draw_text_in_box(email, x + inner_margin, contact_y - (contact_box_height / 2) - 4, email_box_width, field_key='email')
        
        # Second box: phone
        phone = data_dict.get('phone', '')
        phone = transform_phone_number(phone)
        draw_text_in_box(phone, x + inner_margin + email_box_width + contact_box_gap, contact_y - (contact_box_height / 2) - 4, phone_box_width, field_key='phone')
        
        # Third box: profession
        profession = data_dict.get('profession', '')
        draw_text_in_box(profession, x + inner_margin + email_box_width + contact_box_gap + phone_box_width + contact_box_gap, contact_y - (contact_box_height / 2) - 4, profession_box_width, field_key='profession')
        
        # Start a new outer box for legal guardian information
        guardian_y = contact_y - contact_box_height - 30  # Position below contact boxes with 30pt gap (reduced from 50pt)
        
        # Calculate dimensions for the guardian box
        guardian_box_height = 160  # Enlarged height to accommodate all content and extend to bottom
        guardian_box_width = usable_width
        
        # Draw the guardian box
        c.setStrokeColor(black)
        c.setLineWidth(1)
        c.rect(x, guardian_y - guardian_box_height, guardian_box_width, guardian_box_height)
        
        # Add the heading (without underline)
        c.setFont("Helvetica-Bold", 12)
        c.drawString(x + 10, guardian_y - 20, "Bei Minderjährigen: Angaben des gesetzlichen Vertreters:")
        
        # Add gender selection row (left-aligned, indented by 1.5cm)
        c.setFont("Helvetica", 10)
        c.setFillColor(black)
        
        # Get applicant gender value from data
        applicant_sex = data_dict.get('applicant_sex', '')
        if applicant_sex is not None:
            applicant_sex = applicant_sex.lower()
        else:
            applicant_sex = ''
        
        # Create gender selection text with checkboxes
        guardian_gender_text = "Geschlecht: O weiblich / O männlich / O divers"
        
        # Position the text (left-aligned, indented by 1.5cm)
        indent_x = x + 1.5 * cm
        c.drawString(indent_x, guardian_y - 45, guardian_gender_text)
        
        # Add checkmarks by overprinting "x" over "o" based on applicant_sex value
        if applicant_sex in ['weiblich', 'männlich', 'divers']:
            # Calculate positions for the "O" characters
            o_positions = []
            current_x = indent_x
            for i, char in enumerate(guardian_gender_text):
                if char == 'O':
                    o_positions.append(current_x)
                current_x += c.stringWidth(char, "Helvetica", 10)
            
            # Overprint "x" on the appropriate "O"
            if applicant_sex == 'weiblich' and len(o_positions) >= 1:
                c.drawString(o_positions[0], guardian_y - 45, "X")
            elif applicant_sex == 'männlich' and len(o_positions) >= 2:
                c.drawString(o_positions[1], guardian_y - 45, "X")
            elif applicant_sex == 'divers' and len(o_positions) >= 3:
                c.drawString(o_positions[2], guardian_y - 45, "X")
        
        # Add applicant name row (full width)
        applicant_name_y = guardian_y - 70  # Position below gender selection
        
        # Calculate dimensions for applicant name box
        applicant_name_box_height = 1.2 * cm
        applicant_name_box_width = guardian_box_width - 20  # Full width minus margins
        
        # Draw the applicant name box with light grey fill
        c.setFillColor(light_grey)
        c.rect(x + 10, applicant_name_y - applicant_name_box_height, applicant_name_box_width, applicant_name_box_height, fill=1)
        c.setFillColor(black)  # Reset to black for text
        
        # Add label in 8pt font in lower left corner
        c.setFont("Helvetica", 8)
        c.drawString(x + 12, applicant_name_y - applicant_name_box_height + 2, "Name, Vorname (gesetzl. Vertreter)")
        
        # Add content in 12pt font, left-aligned, baseline in middle of box
        applicant_last_name = data_dict.get('applicant_last_name', '')
        applicant_first_name = data_dict.get('applicant_first_name', '')
        applicant_name_text = f"{applicant_last_name}, {applicant_first_name}" if applicant_last_name and applicant_first_name else f"{applicant_last_name}{applicant_first_name}"
        draw_text_in_box(applicant_name_text, x + 10, applicant_name_y - (applicant_name_box_height / 2) - 4, applicant_name_box_width, padding=5, field_key='applicant_last_name')
        
        # Add applicant address row (same layout as member address)
        applicant_address_y = applicant_name_y - applicant_name_box_height - 10  # Position below name box
        
        # Calculate dimensions for applicant address boxes
        applicant_address_box_height = 1.2 * cm
        applicant_address_box_gap = 5  # Gap between applicant address boxes
        applicant_address_box1_width = (applicant_name_box_width * 0.6) - applicant_address_box_gap  # 60% of width minus gap
        applicant_address_box2_width = (applicant_name_box_width * 0.15) - applicant_address_box_gap  # 15% of width minus gap
        applicant_address_box3_width = (applicant_name_box_width * 0.25)  # 25% of width minus gap plus fixed amount
        
        # Draw the three applicant address boxes with light grey fill
        c.setFillColor(light_grey)
        c.rect(x + 10, applicant_address_y - applicant_address_box_height, applicant_address_box1_width, applicant_address_box_height, fill=1)  # Street
        c.rect(x + 10 + applicant_address_box1_width + applicant_address_box_gap, applicant_address_y - applicant_address_box_height, applicant_address_box2_width, applicant_address_box_height, fill=1)  # ZIP
        c.rect(x + 10 + applicant_address_box1_width + applicant_address_box_gap + applicant_address_box2_width + applicant_address_box_gap, applicant_address_y - applicant_address_box_height, applicant_address_box3_width, applicant_address_box_height, fill=1)  # City
        c.setFillColor(black)  # Reset to black for text
        
        # Add labels in 8pt font in lower left corner of each box
        c.setFont("Helvetica", 8)
        c.drawString(x + 12, applicant_address_y - applicant_address_box_height + 2, "Straße, Hausnummer (sofern abweichend vom Mitglied)")
        c.drawString(x + 12 + applicant_address_box1_width + applicant_address_box_gap, applicant_address_y - applicant_address_box_height + 2, "Postleitzahl")
        c.drawString(x + 12 + applicant_address_box1_width + applicant_address_box_gap + applicant_address_box2_width + applicant_address_box_gap, applicant_address_y - applicant_address_box_height + 2, "Ort")
        
        # Add content in 12pt font, left-aligned, baseline in middle of boxes
        c.setFont("Helvetica", 12)
        
        # First box: applicant_street
        applicant_street = data_dict.get('applicant_street', '')
        draw_text_in_box(applicant_street, x + 10, applicant_address_y - (applicant_address_box_height / 2) - 4, applicant_address_box1_width, padding=5, field_key='applicant_street')
        
        # Second box: applicant_zip
        applicant_zip = data_dict.get('applicant_zip', '')
        draw_text_in_box(applicant_zip, x + 10 + applicant_address_box1_width + applicant_address_box_gap, applicant_address_y - (applicant_address_box_height / 2) - 4, applicant_address_box2_width, padding=5, field_key='applicant_zip')
        
        # Third box: applicant_city
        applicant_city = data_dict.get('applicant_city', '')
        draw_text_in_box(applicant_city, x + 10 + applicant_address_box1_width + applicant_address_box_gap + applicant_address_box2_width + applicant_address_box_gap, applicant_address_y - (applicant_address_box_height / 2) - 4, applicant_address_box3_width, padding=5, field_key='applicant_city')
        
        # Add section header for SEPA mandate
        sepa_y = applicant_address_y - applicant_address_box_height - 50  # Position below guardian box with 50pt gap (increased from 30pt)
        
        # Section heading - bold, 12pt, underlined (same as first header)
        c.setFont("Helvetica-Bold", 12)
        sepa_section_text = "II. Erteilung eines SEPA-Lastschriftmandats:"
        c.drawString(x, sepa_y, sepa_section_text)
        
        # Draw underline for section heading
        sepa_text_width = c.stringWidth(sepa_section_text, "Helvetica-Bold", 12)
        c.line(x, sepa_y - 2, x + sepa_text_width, sepa_y - 2)
        
        # Start the SEPA mandate box
        sepa_box_y = sepa_y - 15  # Position below section header (reduced from 30pt)
        
        # Calculate dimensions for the SEPA box
        sepa_box_height = 120  # Enlarged height to properly enclose inner fields
        sepa_box_width = usable_width
        
        # Draw the SEPA box
        c.setStrokeColor(black)
        c.setLineWidth(1)
        c.rect(x, sepa_box_y - sepa_box_height, sepa_box_width, sepa_box_height)
        
        # First row: Account holder (full width)
        account_holder_y = sepa_box_y - 20  # Position inside the box
        
        # Calculate dimensions for account holder box
        account_holder_box_height = 1.2 * cm
        account_holder_box_width = inner_width  # Use same width as other sections
        
        # Draw the account holder box with light grey fill
        c.setFillColor(light_grey)
        c.rect(x + inner_margin, account_holder_y - account_holder_box_height, account_holder_box_width, account_holder_box_height, fill=1)
        c.setFillColor(black)  # Reset to black for text
        
        # Add label in 8pt font in lower left corner
        c.setFont("Helvetica", 8)
        c.drawString(x + inner_margin + 2, account_holder_y - account_holder_box_height + 2, "Name des Kontoinhabers")
        
        # Add content in 12pt font, left-aligned, baseline in middle of box
        account_holder = data_dict.get('account_holder', '')
        draw_text_in_box(account_holder, x + inner_margin, account_holder_y - (account_holder_box_height / 2) - 4, account_holder_box_width, field_key='account_holder')
        
        # Second row: IBAN and BIC (2/3 and 1/3 width)
        iban_bic_y = account_holder_y - account_holder_box_height - 10  # Position below account holder box
        
        # Calculate dimensions for IBAN and BIC boxes
        iban_bic_box_height = 1.2 * cm
        iban_bic_gap = 5  # Gap between IBAN and BIC boxes
        iban_box_width = (account_holder_box_width * 2) / 3 - iban_bic_gap / 2  # 2/3 of width minus gap
        bic_box_width = account_holder_box_width / 3 - iban_bic_gap / 2  # 1/3 of width minus gap
        
        # Draw the IBAN and BIC boxes with light grey fill
        c.setFillColor(light_grey)
        c.rect(x + 10, iban_bic_y - iban_bic_box_height, iban_box_width, iban_bic_box_height, fill=1)  # IBAN
        c.rect(x + 10 + iban_box_width + iban_bic_gap, iban_bic_y - iban_bic_box_height, bic_box_width, iban_bic_box_height, fill=1)  # BIC
        c.setFillColor(black)  # Reset to black for text
        
        # Add labels in 8pt font in lower left corner of each box
        c.setFont("Helvetica", 8)
        c.drawString(x + 12, iban_bic_y - iban_bic_box_height + 2, "IBAN")
        c.drawString(x + 12 + iban_box_width + iban_bic_gap, iban_bic_y - iban_bic_box_height + 2, "BIC")
        
        # Add content in 12pt font, left-aligned, baseline in middle of boxes
        c.setFont("Helvetica", 12)
        
        # IBAN box with formatting (normalize so spaces in source do not break groups)
        iban = data_dict.get('iban', '')
        normalized_iban = normalize_iban(iban)
        if normalized_iban:
            formatted_iban = ' '.join([normalized_iban[i:i + 4] for i in range(0, len(normalized_iban), 4)])
            display_iban = ('!!Ungültig!! ' + formatted_iban) if mark_iban_invalid else formatted_iban
            draw_text_in_box(display_iban, x + 10, iban_bic_y - (iban_bic_box_height / 2) - 4, iban_box_width, padding=5, field_key='iban')
        
        # BIC box (normalize whitespace for display)
        bic = data_dict.get('bic', '')
        normalized_bic = normalize_bic(bic)
        if normalized_bic:
            display_bic = ('!!Ungültig!! ' + normalized_bic) if mark_bic_invalid else normalized_bic
            draw_text_in_box(
                display_bic,
                x + 10 + iban_box_width + iban_bic_gap,
                iban_bic_y - (iban_bic_box_height / 2) - 4,
                bic_box_width,
                padding=5,
                field_key='bic',
            )
        
    except Exception as e:
        raise Exception(f"Error creating PDF page: {e}")


def convert_enrollment_file_to_pdf(
    input_path: str,
    pdf_path: str,
    debug: bool = False,
    validation_issues: Optional[List[str]] = None,
    iban_validation_issues: Optional[List[str]] = None,
) -> None:
    """
    Read a CSV or XLSX enrollment file and convert all rows to a multi-page PDF.

    Args:
        input_path: Path to the input file (.csv or .xlsx)
        pdf_path: Path where the output PDF should be saved
        debug: If True, show source column names as labels in top-right corner of fields
        validation_issues: If provided, XLSX rows append one line per issue:
            \"IBAN: <holder>: <message>\" or \"BIC: <holder>: <message>\".
        iban_validation_issues: Deprecated alias for validation_issues (if both passed, validation_issues wins).

    Raises:
        Exception: If there's an error reading the input or creating the PDF
        ValueError: If the file extension is not supported
    """
    suffix = Path(input_path).suffix.lower()
    if suffix == '.csv':
        data_list = read_csv_to_dicts_with_validation(input_path)
        debug_rev: Optional[Dict[str, str]] = None
        validate_ibans_for_input = False
    elif suffix == '.xlsx':
        data_list = read_xlsx_to_mapped_dicts(input_path)
        debug_rev = xlsx_reverse_key_mapping
        validate_ibans_for_input = True
    else:
        raise ValueError(
            f"Unsupported file type '{suffix}'. Use .csv or .xlsx (legacy .xls is not supported)."
        )

    issues_sink = validation_issues if validation_issues is not None else iban_validation_issues

    try:
        if not data_list:
            raise Exception(f"Input file contains no data rows: {input_path}")

        c = canvas.Canvas(pdf_path, pagesize=A4)

        for i, row in enumerate(data_list):
            mark_iban_invalid = False
            mark_bic_invalid = False
            if validate_ibans_for_input:
                holder = row.get('account_holder') or '(no account holder)'
                iban_raw = row.get('iban', '')
                if normalize_iban(iban_raw):
                    err = validate_iban(iban_raw)
                    if err is not None:
                        mark_iban_invalid = True
                        if issues_sink is not None:
                            issues_sink.append(f'IBAN: {holder}: {err}')
                bic_raw = row.get('bic', '')
                if normalize_bic(bic_raw):
                    err_b = validate_bic(bic_raw)
                    if err_b is not None:
                        mark_bic_invalid = True
                        if issues_sink is not None:
                            issues_sink.append(f'BIC: {holder}: {err_b}')
            generate_pdf_from_dict(
                row,
                c,
                debug,
                debug_reverse_mapping=debug_rev,
                mark_iban_invalid=mark_iban_invalid,
                mark_bic_invalid=mark_bic_invalid,
            )
            if i < len(data_list) - 1:
                c.showPage()

        c.save()

        print(f"✅ Successfully converted {len(data_list)} rows from '{input_path}' to '{pdf_path}'")

    except Exception as e:
        raise Exception(f"Error in conversion process: {e}")


def convert_csv_to_pdf(csv_path: str, pdf_path: str, debug: bool = False) -> None:
    """Backward-compatible alias for CSV-only callers; accepts .xlsx via convert_enrollment_file_to_pdf."""
    convert_enrollment_file_to_pdf(csv_path, pdf_path, debug)



def read_csv_to_dicts_with_validation(filename: str) -> List[Dict[str, Any]]:
    """
    Read a CSV file with cp1252, cp1250, or utf-8 encoding (auto-detected) and semicolon separators,
    with additional validation and error handling.
    
    Args:
        filename (str): Path to the CSV file to read
        
    Returns:
        List[Dict[str, Any]]: List of dictionaries, each representing a row
                              with column headers as keys
                              
    Raises:
        FileNotFoundError: If the specified file doesn't exist
        UnicodeDecodeError: If the file cannot be decoded with supported encodings
        ValueError: If the CSV file is empty, has no headers, or uses an unsupported encoding
        Exception: For other CSV parsing errors
    """
    try:
        # Detect the encoding using charset-normalizer
        detection_result = from_path(filename).best()
        
        if detection_result is None:
            raise ValueError(f"Could not detect encoding for file '{filename}'")
        
        detected_encoding = detection_result.encoding.lower()
        print(f"🔍 Detected encoding: {detected_encoding}")
        
        # Normalize encoding names and validate
        # charset-normalizer may return various names for cp1252, cp1250, and utf-8
        allowed_encodings = {
            'utf-8': 'utf-8',
            'utf_8': 'utf-8',
            'cp1252': 'cp1252',
            'windows-1252': 'cp1252',
            'windows_1252': 'cp1252',
            'iso-8859-1': 'cp1252',  # Often misdetected as ISO-8859-1
            'iso_8859_1': 'cp1252',
            'latin-1': 'cp1252',
            'latin_1': 'cp1252',
            'cp1250': 'cp1250',
            'windows-1250': 'cp1250',
            'windows_1250': 'cp1250',
            'iso-8859-2': 'cp1250',  # Central European encoding
            'iso_8859_2': 'cp1250',
        }
        
        # Map detected encoding to normalized encoding
        normalized_encoding = allowed_encodings.get(detected_encoding)
        
        if normalized_encoding is None:
            raise ValueError(
                f"Unsupported encoding '{detected_encoding}' detected for file '{filename}'. "
                f"Only cp1252, cp1250, and utf-8 encodings are supported."
            )
        
        print(f"✅ Using encoding: {normalized_encoding}")
        
        with open(filename, 'r', encoding=normalized_encoding, newline='') as csvfile:
            # Read the first line to check if file has content
            first_line = csvfile.readline().strip()
            if not first_line:
                raise ValueError(f"File '{filename}' is empty")
            
            # Reset file pointer to beginning
            csvfile.seek(0)
            
            reader = csv.DictReader(csvfile, delimiter=';')
            
            # Check if headers exist
            if not reader.fieldnames:
                raise ValueError(f"File '{filename}' has no headers")
            
            # Convert to list and validate
            data = list(reader)
            
            # Check if we have any data rows
            if not data:
                print(f"Warning: File '{filename}' contains only headers, no data rows")
            
            # Convert the data using the key mapping for all rows
            for i, row in enumerate(data):
                converted_row = {}
                for original_key, mapped_key in key_mapping.items():
                    if mapped_key is not None:
                        converted_row[mapped_key] = row.get(original_key, '')
                data[i] = converted_row
            
            return _filter_nonempty_rows(data)
            
    except FileNotFoundError:
        raise FileNotFoundError(f"File '{filename}' not found")
    except UnicodeDecodeError as e:
        raise UnicodeDecodeError(f"Failed to decode file '{filename}': {e}")
    except Exception as e:
        raise Exception(f"Error reading CSV file '{filename}': {e}")


def main():
    """Main function to handle command line arguments"""
    if len(sys.argv) < 3 or len(sys.argv) > 4:
        print("Usage: python csv_converter.py <input_csv_or_xlsx_path> <output_pdf_path> [--no-debug]")
        print("Example: python csv_converter.py enrollments.csv output.pdf")
        print("Example: python csv_converter.py enrollments.xlsx output.pdf --no-debug")
        sys.exit(1)
    
    csv_path = sys.argv[1]
    pdf_path = sys.argv[2]
    debug = True if len(sys.argv) == 3 or sys.argv[3] != "--no-debug" else False
    
    try:
        convert_enrollment_file_to_pdf(csv_path, pdf_path, debug)
    except Exception as e:
        print(f"❌ Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()